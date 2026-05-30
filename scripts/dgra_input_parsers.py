#!/usr/bin/env python3
"""
DGRA Input Parsers — Unified input layer for v0.5 P0-1
Supports VCF (.vcf/.vcf.gz), Excel (.xlsx), TSV/CSV, and free text.

用法:
    from dgra_input_parsers import auto_detect, VCFParser, ExcelParser, TSVParser, FreeTextParser
    fmt = auto_detect(Path("donor.vcf.gz"))  # -> "vcf"
    parser = VCFParser()
    variants = parser.parse(Path("donor.vcf.gz"))
"""

import re
import csv
import gzip
from pathlib import Path
from typing import List, Dict, Any, Optional

# =============================================================================
# VEP CSQ column mapping (canonical order from VEP default output)
# =============================================================================

# v0.10.0 P2-3: Configurable VEP annotation field name.
# Default "CSQ" works for standard VEP VCF output.
# Use "ANN" for SnpEff or custom VEP --vcf_info_field values.
VEP_ANNOTATION_FIELD: str = "CSQ"

VEP_CSQ_FIELDS = [
    "Allele", "Consequence", "IMPACT", "SYMBOL", "Gene", "Feature",
    "Feature_type", "EXON", "INTRON", "HGVSc", "HGVSp", "cDNA_position",
    "CDS_position", "Protein_position", "Amino_acids", "Codons",
    "Existing_variation", "DISTANCE", "STRAND", "FLAGS", "SYMBOL_SOURCE",
    "HGNC_ID", "CANONICAL", "MANE_SELECT", "MANE_PLUS_CLINICAL", "TSL",
    "APPRIS", "CCDS", "ENSP", "SWISSPROT", "TREMBL", "UNIPARC",
    "UNIPROT_ISOFORM", "GENE_PHENO", "SIFT", "PolyPhen", "DOMAINS", "miRNA",
    "AF", "AFR_AF", "AMR_AF", "EAS_AF", "EUR_AF", "SAS_AF",
    "gnomAD_AF", "gnomAD_AFR_AF", "gnomAD_AMR_AF", "gnomAD_ASJ_AF",
    "gnomAD_EAS_AF", "gnomAD_FIN_AF", "gnomAD_NFE_AF", "gnomAD_OTH_AF",
    "gnomAD_SAS_AF", "MAX_AF", "MAX_AF_POPS", "CLIN_SIG", "SOMATIC",
    "PHENO", "PUBMED", "MOTIF_NAME", "MOTIF_SCORE_CHANGE", "TRANSCRIPTION_FACTORS",
]

VEP_CSQ_MAP = {name: idx for idx, name in enumerate(VEP_CSQ_FIELDS)}

# Map VEP CSQ internal names to dgra_core REQUIRED_COLS
CSQ_TO_DGRA = {
    "SYMBOL": "GENE",
    "Feature": "Feature",
    "EXON": "EXON",
    "IMPACT": "IMPACT",
    "Consequence": "Consequence",
    "HGVSp": "HGVSp",
    "HGVSc": "HGVSc",
    "CLIN_SIG": "CLIN_SIG",
    "gnomAD_AF": "gnomAD_AF",
}

# =============================================================================
# Format auto-detection
# =============================================================================

def auto_detect(path: Path) -> str:
    """Detect input format from extension + file content."""
    suffix = path.suffix.lower()
    name = path.name.lower()

    if suffix == ".gz" or name.endswith(".vcf.gz"):
        return "vcf"
    if suffix in (".vcf", ".bcf"):
        return "vcf"
    if suffix in (".xlsx", ".xlsm"):
        return "excel"
    if suffix in (".tsv", ".csv"):
        # Peek first line to distinguish TSV from CSV
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            first = f.readline()
            if "\t" in first:
                return "tsv"
            return "csv"
    if suffix in (".txt", ".md"):
        return "freetext"

    # Fallback: read first 2KB and look for VCF header
    with open(path, "rb", encoding='utf-8') as f:
        head = f.read(2048)
    if b"##fileformat=VCF" in head:
        return "vcf"
    if b"##INFO=<ID=CSQ" in head:
        return "vcf"

    raise ValueError(f"Cannot auto-detect format for {path}. Use --format to specify.")


# =============================================================================
# Base class
# =============================================================================

class InputParser:
    """Base input parser."""
    def parse(self, path: Path) -> List[Dict[str, Any]]:
        raise NotImplementedError


# =============================================================================
# TSV / CSV Parser (legacy compatible)
# =============================================================================

class TSVParser(InputParser):
    """Parse TSV or CSV files. Assumes header row matches VEP-style columns."""
    def __init__(self, dialect: str = "auto", adapter: Optional[Any] = None):
        self.dialect = dialect  # "auto" | "tab" | "comma"
        self.adapter = adapter  # AnnotationAdapter instance (optional)

    def parse(self, path: Path) -> List[Dict[str, Any]]:
        variants: List[Dict[str, Any]] = []
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            if self.dialect == "auto":
                sample = f.read(8192)
                f.seek(0)
                delimiter = "\t" if "\t" in sample else ","
            else:
                delimiter = "\t" if self.dialect == "tab" else ","
            reader = csv.DictReader(f, delimiter=delimiter)
            # v0.9.1: Detect and translate Chinese headers before parsing
            raw_fieldnames = reader.fieldnames or []
            from gpa_i18n import is_chinese_header, translate_chinese_headers
            if is_chinese_header(raw_fieldnames):
                translated = translate_chinese_headers(raw_fieldnames)
                # Rebuild reader with translated headers
                f.seek(0)
                # Skip header line
                next(f)
                reader = csv.DictReader(f, fieldnames=translated, delimiter=delimiter)
            raw_rows = []
            for row in reader:
                # Strip whitespace, convert empty strings to empty (core.py handles UNKNOWN)
                clean = {k.strip(): (v.strip() if v is not None else "") for k, v in row.items()}
                raw_rows.append(clean)

        # v0.5 P0-2: Apply annotation adapter if provided or auto-detect
        if self.adapter is not None:
            from dgra_adapters import adapt_rows
            variants = adapt_rows(raw_rows, adapter=self.adapter)
        elif raw_rows:
            from dgra_adapters import adapt_rows
            variants = adapt_rows(raw_rows, adapter=None)
        else:
            variants = raw_rows
        return variants


# =============================================================================
# VCF Parser (cyvcf2)
# =============================================================================

class VCFParser(InputParser):
    """
    Parse VCF using pure Python (stdlib only).

    Supports:
      - Plain VCF and uncompressed .vcf.gz (ASCII text)
      - bgzipped VCF via gzip module
      - VEP-annotated VCF (INFO/CSQ)
      - Multi-allelic sites split into separate records
      - GT, DP, GQ, VAF extracted from FORMAT

    v0.10.9: Replaced cyvcf2 (C extension, macOS Team ID signature issue)
    with pure-Python parser that works in any sandboxed environment.
    """
    def __init__(self, sample_idx: int = 0, prefer_canonical: bool = True,
                 keep_all_transcripts: bool = False):
        self.sample_idx = sample_idx
        self.prefer_canonical = prefer_canonical
        self.keep_all_transcripts = keep_all_transcripts

    def _open_vcf(self, path: Path):
        """Open plain or gzip-compressed VCF. Returns iterable of lines."""
        path = Path(path)
        # Try gzip first (handles both real .gz and ASCII .vcf.gz)
        try:
            f = gzip.open(str(path), "rt", encoding="utf-8", errors="replace")
            # Peek to confirm it's readable
            f.read(1)
            f.seek(0)
            return f
        except (OSError, gzip.BadGzipFile):
            return open(str(path), "r", encoding="utf-8", errors="replace")

    def _parse_csq_header_lines(self, header_lines: List[str]) -> Dict[str, int]:
        """Extract CSQ field order from header lines."""
        for line in header_lines:
            if 'ID=CSQ' in line and 'Description=' in line:
                m = re.search(r'Format:\s*([^"]+)"', line)
                if m:
                    fields = m.group(1).strip().split("|")
                    return {name: idx for idx, name in enumerate(fields)}
        return VEP_CSQ_MAP  # fallback

    def _pick_csq(self, csq_entries: List[List[str]], csq_map: Dict[str, int]) -> List[str]:
        """Pick one transcript per allele. Prefer CANONICAL=YES or MANE_SELECT."""
        if not csq_entries:
            return []
        if self.keep_all_transcripts:
            return csq_entries[0]
        for entry in csq_entries:
            can_idx = csq_map.get("CANONICAL")
            if can_idx is not None and len(entry) > can_idx and entry[can_idx] == "YES":
                return entry
        for entry in csq_entries:
            mane_idx = csq_map.get("MANE_SELECT")
            if mane_idx is not None and len(entry) > mane_idx and entry[mane_idx]:
                return entry
        return csq_entries[0]

    def _parse_info(self, info_str: str) -> Dict[str, str]:
        """Parse VCF INFO field into a dict."""
        info = {}
        if info_str == "." or not info_str:
            return info
        for part in info_str.split(";"):
            if "=" in part:
                k, _, v = part.partition("=")
                info[k] = v
            else:
                info[part] = "1"
        return info

    def _parse_format(self, fmt_str: str, sample_str: str) -> Dict[str, str]:
        """Parse FORMAT + sample string into a dict."""
        if not fmt_str or not sample_str:
            return {}
        keys = fmt_str.split(":")
        vals = sample_str.split(":")
        return {k: (vals[i] if i < len(vals) else "") for i, k in enumerate(keys)}

    def _extract_gt_fields(self, fmt_dict: Dict[str, str]):
        """Extract GT, DP, GQ, VAF from parsed FORMAT dict."""
        gt_raw = fmt_dict.get("GT", "").replace("|", "/")
        dp = fmt_dict.get("DP", "")
        gq = fmt_dict.get("GQ", "")
        vaf = ""
        # VAF from AD
        ad_str = fmt_dict.get("AD", "")
        if ad_str and ad_str != ".":
            try:
                parts = [int(x) for x in ad_str.split(",") if x not in (".", "")]
                if len(parts) >= 2:
                    total = sum(parts)
                    if total > 0:
                        alt_depth = sum(parts[1:])
                        vaf = f"{alt_depth / total:.4f}"
            except (RuntimeError, ValueError):
                pass
        if not vaf:
            af_str = fmt_dict.get("AF", "")
            if af_str and af_str not in (".", ""):
                vaf = af_str.split(",")[0]
        return gt_raw, dp, gq, vaf

    def parse(self, path: Path) -> List[Dict[str, Any]]:
        variants: List[Dict[str, Any]] = []
        header_lines = []
        col_names = []
        csq_map: Dict[str, int] = {}
        samples: List[str] = []

        fh = self._open_vcf(path)
        try:
            for raw_line in fh:
                line = raw_line.rstrip("\n\r")
                if line.startswith("##"):
                    header_lines.append(line)
                    continue
                if line.startswith("#CHROM"):
                    col_names = line.lstrip("#").split("\t")
                    # samples start at column index 9
                    samples = col_names[9:] if len(col_names) > 9 else []
                    csq_map = self._parse_csq_header_lines(header_lines)
                    continue
                if not line or not col_names:
                    continue

                fields = line.split("\t")
                if len(fields) < 8:
                    continue

                chrom = fields[0].replace("chr", "")
                pos = int(fields[1])
                ref = fields[3]
                alt_field = fields[4]
                info_str = fields[7] if len(fields) > 7 else ""
                fmt_str = fields[8] if len(fields) > 8 else ""
                sample_str = fields[9] if len(fields) > 9 else ""

                # Handle multi-allelic ALTs
                alts = [a for a in alt_field.split(",") if a != "."]
                if not alts:
                    continue

                fmt_dict = self._parse_format(fmt_str, sample_str)
                gt_raw, dp, gq, vaf = self._extract_gt_fields(fmt_dict)

                info = self._parse_info(info_str)
                csq_raw = info.get(VEP_ANNOTATION_FIELD, "")

                if csq_raw:
                    csq_strings = csq_raw.split(",")
                    for alt in alts:
                        allele_csq = []
                        for csq_str in csq_strings:
                            parts = csq_str.split("|")
                            if parts[0] == alt:
                                allele_csq.append(parts)
                        chosen = self._pick_csq(allele_csq, csq_map)
                        variant = self._csq_to_variant(
                            chrom, pos, ref, alt, chosen, csq_map,
                            gt=gt_raw, dp=dp, gq=gq, vaf=vaf
                        )
                        variants.append(variant)
                else:
                    # No VEP annotation — emit minimal record
                    for alt in alts:
                        variants.append({
                            "CHROM": chrom,
                            "POS": str(pos),
                            "REF": ref,
                            "ALT": alt,
                            "GENE": "",
                            "Feature": "",
                            "EXON": "",
                            "IMPACT": "",
                            "Consequence": "",
                            "HGVSp": "",
                            "HGVSc": "",
                            "CLIN_SIG": "",
                            "GT": gt_raw,
                            "DP": dp,
                            "GQ": gq,
                            "VAF": vaf,
                            "gnomAD_AF": "",
                        })
        finally:
            fh.close()
        return variants

    def _csq_to_variant(self, chrom: str, pos: int, ref: str, alt: str,
                        csq: List[str], csq_map: Dict[str, int],
                        gt: str, dp: str, gq: str, vaf: str) -> Dict[str, Any]:
        """Map one CSQ entry to dgra_core REQUIRED_COLS."""
        def get(field: str) -> str:
            idx = csq_map.get(field)
            if idx is None or idx >= len(csq):
                return ""
            val = csq[idx]
            return val if val else ""

        # gnomAD_AF may contain multiple values (e.g., "0.001&0.002")
        gnomad_raw = get("gnomAD_AF")
        if "&" in gnomad_raw:
            gnomad_raw = gnomad_raw.split("&")[0]

        return {
            "CHROM": chrom,
            "POS": str(pos),
            "REF": ref,
            "ALT": alt,
            "GENE": get("SYMBOL"),
            "Feature": get("Feature"),
            "EXON": get("EXON"),
            "IMPACT": get("IMPACT"),
            "Consequence": get("Consequence"),
            "HGVSp": get("HGVSp"),
            "HGVSc": get("HGVSc"),
            "CLIN_SIG": get("CLIN_SIG"),
            "GT": gt,
            "DP": dp,
            "GQ": gq,
            "VAF": vaf,
            "gnomAD_AF": gnomad_raw,
        }


# =============================================================================
# Excel Parser (openpyxl)
# =============================================================================

class ExcelParser(InputParser):
    """
    Parse Excel .xlsx files.

    Auto-detects:
      - Active or first sheet
      - Header row (first non-empty row)
      - Column name fuzzy matching for VEP/ANNOVAR/SnpEff (basic)

    P0-1: VEP-style columns are primary; ANNOVAR/SnpEff mapping is minimal.
    Full adapter layer (P0-2) will handle detailed column remapping.
    """
    def __init__(self, sheet_name: Optional[str] = None, header_row: int = 1):
        self.sheet_name = sheet_name
        self.header_row = header_row

    def _normalize_header(self, cell_value: Any) -> str:
        """Strip, upper-case, remove special chars."""
        if cell_value is None:
            return ""
        s = str(cell_value).strip()
        s = s.replace("#", "").replace(" ", "_")
        return s

    def _fuzzy_col_map(self, headers: List[str]) -> Dict[str, str]:
        """Map header names to dgra_core REQUIRED_COLS."""
        # Direct matches (case-insensitive)
        direct = {
            "chrom": "CHROM", "chr": "CHROM", "#chrom": "CHROM", "chromosome": "CHROM",
            "pos": "POS", "position": "POS", "start": "POS",
            "ref": "REF", "reference": "REF",
            "alt": "ALT", "alternate": "ALT", "obs": "ALT",
            "gene": "GENE", "symbol": "GENE", "gene_name": "GENE", "genename": "GENE",
            "feature": "Feature", "transcript": "Feature", "tx": "Feature",
            "exon": "EXON",
            "impact": "IMPACT",
            "consequence": "Consequence", "anno": "Consequence",
            "hgvsp": "HGVSp", "aa_change": "HGVSp", "protein_change": "HGVSp",
            "hgvsc": "HGVSc", "cdna_change": "HGVSc",
            "clin_sig": "CLIN_SIG", "clinvar": "CLIN_SIG",
            "gt": "GT", "genotype": "GT",
            "dp": "DP", "depth": "DP",
            "gq": "GQ", "quality": "GQ",
            "vaf": "VAF", "af": "VAF", "allele_freq": "VAF",
            "gnomad_af": "gnomAD_AF", "af_gnomad": "gnomAD_AF", "gnomad": "gnomAD_AF",
        }
        mapping: Dict[str, str] = {}
        for h in headers:
            key = h.lower().strip()
            if key in direct:
                mapping[direct[key]] = h
        return mapping

    def parse(self, path: Path) -> List[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("openpyxl is required for Excel parsing. Install: pip install openpyxl") from e

        wb = openpyxl.load_workbook(str(path), data_only=True)
        if self.sheet_name:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active or wb.worksheets[0]

        # Find header row
        headers: List[str] = []
        header_idx = self.header_row
        for row in ws.iter_rows(min_row=header_idx, max_row=header_idx + 5, values_only=True):
            candidate = [self._normalize_header(c) for c in row]
            if any(candidate):
                headers = candidate
                break
        if not headers:
            raise ValueError(f"Could not find header row in {path}")

        col_map = self._fuzzy_col_map(headers)
        variants: List[Dict[str, Any]] = []

        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            # Skip completely empty rows
            if not any(v is not None and str(v).strip() for v in row):
                continue
            variant: Dict[str, Any] = {}
            for dgra_col in [
                "CHROM", "POS", "REF", "ALT", "GENE", "Feature", "EXON",
                "IMPACT", "Consequence", "HGVSp", "HGVSc", "CLIN_SIG",
                "GT", "DP", "GQ", "VAF", "gnomAD_AF",
            ]:
                src = col_map.get(dgra_col)
                if src is not None:
                    idx = headers.index(src)
                    val = row[idx] if idx < len(row) else None
                    variant[dgra_col] = str(val) if val is not None else ""
                else:
                    variant[dgra_col] = ""
            variants.append(variant)

        wb.close()
        return variants


# =============================================================================
# Free Text Parser
# =============================================================================

class FreeTextParser:
    """
    Parse free-text variant descriptions into variant dicts.

    Supported patterns:
      ① c.HGVS:  "TP53 c.722C>T"  → gene + HGVSc
      ② Genomic: "chr17:7578406C>A" or "17-7578406-C-A" → coordinates
      ③ p.HGVS:  "TP53 p.Arg249Ser" → gene + HGVSp

    Offline mode: coordinates may be absent (c./p. HGVS only).
    They are preserved in HGVSc/HGVSp and left for core.py to assess
    via gene-based rules even without exact coordinates.
    """
    # Pattern ①: Gene c.HGVS  e.g. "TP53 c.722C>T", "BRCA1 c.68_69delAG"
    RE_C_HGVS = re.compile(
        r"^\s*([A-Z0-9]+)\s+"
        r"c\.([0-9_+-]+(?:[ACGT]>[ACGT]|del[ACGT]+|ins[ACGT]+|dup[ACGT]+))\s*$",
        re.IGNORECASE,
    )
    # Pattern ②a: chr:posRef>Alt  e.g. "chr17:7578406C>A"
    RE_COORD = re.compile(
        r"^\s*(?:chr)?([0-9XYM]+)[:\-]([0-9]+)\s*([ACGT]+)[>/:]?([ACGT]+)\s*$",
        re.IGNORECASE,
    )
    # Pattern ②b: chr pos ref alt  e.g. "17 7578406 C A"
    RE_COORD_SPACE = re.compile(
        r"^\s*(?:chr)?([0-9XYM]+)\s+([0-9]+)\s+([ACGT]+)\s+([ACGT]+)\s*$",
        re.IGNORECASE,
    )
    # Pattern ③: Gene p.HGVS  e.g. "TP53 p.Arg249Ser", "TP53 p.R249S"
    RE_P_HGVS = re.compile(
        r"^\s*([A-Z0-9]+)\s+"
        r"p\.([A-Za-z*0-9]+)\s*$",
        re.IGNORECASE,
    )

    def parse_text(self, text: str) -> List[Dict[str, Any]]:
        """Parse a single free-text line."""
        text = text.strip()
        if not text:
            return []

        # ②a coordinate
        m = self.RE_COORD.match(text)
        if m:
            chrom, pos, ref, alt = m.groups()
            return [{
                "CHROM": chrom,
                "POS": pos,
                "REF": ref.upper(),
                "ALT": alt.upper(),
                "GENE": "",
                "Feature": "",
                "EXON": "",
                "IMPACT": "",
                "Consequence": "",
                "HGVSp": "",
                "HGVSc": "",
                "CLIN_SIG": "",
                "GT": "",
                "DP": "",
                "GQ": "",
                "VAF": "",
                "gnomAD_AF": "",
            }]

        # ②b coordinate (space-separated)
        m = self.RE_COORD_SPACE.match(text)
        if m:
            chrom, pos, ref, alt = m.groups()
            return [{
                "CHROM": chrom,
                "POS": pos,
                "REF": ref.upper(),
                "ALT": alt.upper(),
                "GENE": "",
                "Feature": "",
                "EXON": "",
                "IMPACT": "",
                "Consequence": "",
                "HGVSp": "",
                "HGVSc": "",
                "CLIN_SIG": "",
                "GT": "",
                "DP": "",
                "GQ": "",
                "VAF": "",
                "gnomAD_AF": "",
            }]

        # ① c.HGVS
        m = self.RE_C_HGVS.match(text)
        if m:
            gene, hgvsc = m.groups()
            return [{
                "CHROM": "",
                "POS": "",
                "REF": "",
                "ALT": "",
                "GENE": gene.upper(),
                "Feature": "",
                "EXON": "",
                "IMPACT": "",
                "Consequence": "",
                "HGVSp": "",
                "HGVSc": f"c.{hgvsc}",
                "CLIN_SIG": "",
                "GT": "",
                "DP": "",
                "GQ": "",
                "VAF": "",
                "gnomAD_AF": "",
            }]

        # ③ p.HGVS
        m = self.RE_P_HGVS.match(text)
        if m:
            gene, hgvsp = m.groups()
            return [{
                "CHROM": "",
                "POS": "",
                "REF": "",
                "ALT": "",
                "GENE": gene.upper(),
                "Feature": "",
                "EXON": "",
                "IMPACT": "",
                "Consequence": "",
                "HGVSp": f"p.{hgvsp}",
                "HGVSc": "",
                "CLIN_SIG": "",
                "GT": "",
                "DP": "",
                "GQ": "",
                "VAF": "",
                "gnomAD_AF": "",
            }]

        raise ValueError(f"Cannot parse free-text variant: '{text}'")

    def parse(self, path: Path) -> List[Dict[str, Any]]:
        """Parse a text file with one variant per line (ignores blank/comment lines)."""
        variants: List[Dict[str, Any]] = []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                variants.extend(self.parse_text(line))
        return variants


# =============================================================================
# Convenience: unified dispatch
# =============================================================================

def parse_input(path: Path, fmt: Optional[str] = None, annotation_fmt: Optional[str] = None) -> List[Dict[str, Any]]:
    """Auto-detect or use specified format, return list of variant dicts.
    v0.5 P0-2: annotation_fmt selects the annotation adapter for TSV/CSV files.
    """
    if fmt is None or fmt == "auto":
        fmt = auto_detect(path)

    # P0-2: Build annotation adapter for TSV/CSV if requested
    adapter = None
    if fmt in ("tsv", "csv") and annotation_fmt and annotation_fmt != "auto":
        from dgra_adapters import VEPAdapter, ANNOVARAdapter, SnpEffAdapter
        adapter_map = {
            "vep": VEPAdapter(),
            "annovar": ANNOVARAdapter(),
            "snpeff": SnpEffAdapter(),
        }
        adapter = adapter_map.get(annotation_fmt)

    if fmt in ("tsv", "csv"):
        return TSVParser(dialect="tab" if fmt == "tsv" else "comma", adapter=adapter).parse(path)
    elif fmt == "vcf":
        result = VCFParser().parse(path)
        # v0.10.3: Warn if raw VCF (no CSQ) was parsed — downstream pipeline
        # will silently downgrade all variants to Tier 3 without annotation.
        if result and len(result) > 0:
            n_empty = sum(
                1 for v in result
                if not v.get("GENE") and not v.get("IMPACT") and not v.get("Consequence")
            )
            if n_empty == len(result):
                import warnings
                warnings.warn(
                    f"VCF '{path}' has no VEP/CSQ annotation. "
                    f"All {len(result)} variants parsed with empty Gene/IMPACT/Consequence. "
                    f"Raw VCF must be annotated before pipeline entry. "
                    f"Use VCFAnnotator, run_gpa_from_file(), or dgra_core.py --input for raw VCFs.",
                    UserWarning,
                    stacklevel=2,
                )
        return result
    elif fmt == "excel":
        return ExcelParser().parse(path)
    elif fmt == "freetext":
        return FreeTextParser().parse(path)
    else:
        raise ValueError(f"Unsupported format: {fmt}")
